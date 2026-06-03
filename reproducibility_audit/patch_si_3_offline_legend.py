#!/usr/bin/env python3
"""Insert explicit OFFLINE definition into SI_3_ReproducibilityAudit.xlsx legend.

Author: Paola Vera-Licona (Vera-Licona Research Group,
University of Connecticut School of Medicine)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX = REPO_ROOT / "data" / "SI_3_ReproducibilityAudit.xlsx"
SHEET = "S2 Table"

# Insert after "• Tool availability:" (row 69).
INSERT_AT = 70
N_ROWS = 10

SECTION_FONT = Font(bold=True, size=11, color="C00000")
BODY_FONT = Font(size=11, color="000000")
WRAP = Alignment(vertical="top", wrap_text=True)

# (label, detail, merge_label_across_row)
BLOCKS: list[tuple[str, str | None, bool]] = [
    ("Tool Avail (column E):", None, True),
    (
        "• Yes:",
        "Public source repository and/or a web tool that loaded at audit.",
        False,
    ),
    (
        "• No:",
        "No publicly accessible source code or web tool.",
        False,
    ),
    (
        "• ⚠️OFFLINE:",
        "Web-only access path; the paper's URL did not load at audit (definition below). "
        "Repro Score and FAIR4RS Score are N/A.",
        False,
    ),
    (
        "OFFLINE — operational definition (December 2025 audit)",
        None,
        True,
    ),
    (
        "During the original 51-method reproducibility audit (repository snapshot "
        "10–11 December 2025), we tested each method's published access path the way a "
        "reader would: open the URL given in the primary paper (desktop browser or "
        "equivalent direct HTTP request). OFFLINE means that interface did not become "
        "usable within the session (connection timed out, host unreachable, or equivalent "
        "failure). We did not treat company marketing pages, alternate host variants "
        "(e.g. apex vs www when the paper lists www), or availability after the audit "
        "window as evidence that the published web tool worked at audit.",
        None,
        True,
    ),
    (
        "Methods marked ⚠️OFFLINE in this table",
        None,
        True,
    ),
    (
        "• Mogrify (PMID 26780608):",
        "http://www.mogrify.net — did not load (connection timed out), December 2025.",
        False,
    ),
    (
        "• PC3T (PMID 37758874):",
        "http://pc3t.idrug.net.cn — did not load (connection timed out), December 2025.",
        False,
    ),
    (
        "Note: DECCODE (row 31) is Tool Avail = Yes because source code is archived on Zenodo; "
        "its companion web UI (https://fantom.gsc.riken.jp/5/cellconv/) was also unreachable in "
        "December 2025 and is documented in Supporting Information S1 (Code Availability), not as "
        "⚠️OFFLINE in column E.",
        None,
        True,
    ),
]


def _merge_row(ws, row: int, col_start: int = 1, col_end: int = 15) -> None:
    ws.merge_cells(
        start_row=row,
        start_column=col_start,
        end_row=row,
        end_column=col_end,
    )


def _unmerge_row(ws, row: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row:
            try:
                ws.unmerge_cells(str(rng))
            except KeyError:
                pass


def _already_patched(ws) -> bool:
    for row in range(INSERT_AT, min(INSERT_AT + 25, ws.max_row + 1)):
        val = ws.cell(row, 1).value
        if val and "OFFLINE — operational definition" in str(val):
            return True
    return False


def patch() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    if _already_patched(ws):
        print(f"SKIP: OFFLINE legend already present in {XLSX.name}")
        return

    _unmerge_row(ws, INSERT_AT)
    ws.insert_rows(INSERT_AT, N_ROWS)

    for offset, (label, detail, merge_label) in enumerate(BLOCKS):
        row = INSERT_AT + offset
        _unmerge_row(ws, row)
        cell_a = ws.cell(row, 1, label)
        cell_a.font = SECTION_FONT if merge_label else BODY_FONT
        cell_a.alignment = WRAP
        if merge_label:
            _merge_row(ws, row)
        elif detail:
            _merge_row(ws, row, col_start=2, col_end=15)
            cell_b = ws.cell(row, 2, detail)
            cell_b.font = BODY_FONT
            cell_b.alignment = WRAP
        ws.row_dimensions[row].height = 36 if merge_label and not detail else 22

    wb.save(XLSX)
    print(f"Updated {XLSX} — inserted OFFLINE legend at row {INSERT_AT}")


if __name__ == "__main__":
    patch()
