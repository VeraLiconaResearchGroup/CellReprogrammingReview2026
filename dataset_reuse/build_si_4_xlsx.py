#!/usr/bin/env python3
"""Build SI_4_DatasetReuse.xlsx — styled like S1/S2 submission tables.

Author: Paola Vera-Licona (Vera-Licona Research Group,
University of Connecticut School of Medicine)
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
DATA = HERE / "data"
OUT = HERE / "SI_4_DatasetReuse.xlsx"
OUT_LEGACY = HERE / "SI_4_DatasetReuse_legacy58.xlsx"

ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Match Table_S1 / Table_S2 submission styling
TITLE_FONT = Font(bold=True, size=12, color="C00000")
META_FONT = Font(size=9, color="C00000")
SECTION_FONT = Font(bold=True, size=11, color="C00000")
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
USED_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

TITLE = (
    "Supporting Information S4 (dataset reuse). Validation-dataset catalog and method × dataset "
    "reuse matrices (v2, 57 methods)"
)
META = (
    "Generated: 2026-05-27 (v2). Corpus: 57 computational cell reprogramming methods "
    "(2013–2026). Source CSVs: dataset_reuse/data/ in this repository. "
    "Main text: cite as Supporting Information S4 (dataset reuse) (§5)."
)
MATRIX_META = (
    "Values: 1 = method cites this resource in validation / data availability; 0 = not catalogued. "
    "Light blue = used. Column labels: Zenodo records show as “DOI 10.5281/zenodo.…” (resolve at "
    "https://doi.org/…). Not every column is raw omics — see sheet S3_accession_key and the note "
    "on S3 Table."
)
ZENODO_CANON = re.compile(r"^10\.5281/ZENODO\.(\d+)$", re.I)
ZENODO_INLINE = re.compile(r"10\.5281/ZENODO\.(\d+)", re.I)

ZENODO_NOTE = (
    "Zenodo columns (DOI 10.5281/zenodo.…): archival deposits cited in method papers — often "
    "inferred GRNs, processed objects, code/tutorial bundles, or paper-specific benchmark files. "
    "They are not always raw RNA-seq series like GEO GSE* records. Open any DOI at "
    "https://doi.org/<DOI> (e.g. https://doi.org/10.5281/zenodo.3921913)."
)


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.select_dtypes(include=["object"]).columns:
        out[col] = out[col].astype(str).map(lambda x: ILLEGAL.sub("", x) if x != "nan" else "")
    return out


def dataset_type(ds: str) -> str:
    if ds in ("ENCODE", "LINCS_L1000", "CMAP", "FANTOM5"):
        return "Compendium"
    if ds == "GPL570":
        return "GEO platform"
    if ds.startswith("GSE") or ds.startswith("GSM"):
        return "GEO"
    if ds.startswith("SRR") or ds.startswith("SRP") or ds.startswith("SRX"):
        return "SRA"
    if ds.startswith("PRJNA") or ds.startswith("PRJEB"):
        return "BioProject"
    if ds.startswith("E-MTAB"):
        return "ArrayExpress"
    if ZENODO_CANON.match(ds.strip()):
        return "Zenodo (DOI deposit)"
    if ds.isupper() or "_" in ds:
        return "Compendium/Resource"
    return "Other"


def format_accession_display(raw: str) -> str:
    """Human-readable column / cell label (Zenodo → DOI prefix)."""
    if raw == "Method":
        return raw
    text = raw.strip()
    match = ZENODO_CANON.match(text)
    if match:
        return f"DOI 10.5281/zenodo.{match.group(1)}"
    return text


def zenodo_resolve_url(raw: str) -> str:
    match = ZENODO_CANON.match(raw.strip())
    if match:
        return f"https://doi.org/10.5281/zenodo.{match.group(1)}"
    return ""


def format_datasets_used_list(ds: str) -> str:
    if not ds or ds == "nan":
        return ""
    parts = re.split(r"[;,]", ds)
    formatted = [format_accession_display(p.strip()) for p in parts if p.strip()]
    return "; ".join(formatted)


def accession_key_note(raw: str) -> str:
    if ZENODO_CANON.match(raw.strip()):
        return "Zenodo deposit (see S3 Table footer); may be GRN/processed data/code, not raw GEO"
    if raw in ("ENCODE", "LINCS_L1000", "CMAP", "FANTOM5", "GPL570"):
        return "Shared compendium (see S3_shared)"
    return ""


def autosize_columns(ws, min_width: float = 8.0, max_width: float = 48.0) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 120)):
            val = ws.cell(row=row, column=col_idx).value
            if val is None:
                continue
            for line in str(val).splitlines():
                max_len = max(max_len, len(line))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_width), max_width)


def style_title_block(ws, header_row: int) -> None:
    ws.row_dimensions[1].height = 25
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.cell(1, 1, TITLE).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(vertical="center", wrap_text=True)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    ws.cell(2, 1, META).font = META_FONT
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def style_data_rows(ws, start_row: int, end_row: int, ncols: int, alt: bool = True) -> None:
    for i, row in enumerate(range(start_row, end_row + 1)):
        row_fill = (WHITE_FILL if i % 2 == 0 else GRAY_FILL) if alt else WHITE_FILL
        for col in range(1, ncols + 1):
            cell = ws.cell(row, col)
            if cell.fill.fill_type is None or cell.fill.start_color.rgb in ("00000000", "FFFFFFFF", "00FFFFFF", "00F2F2F2"):
                cell.fill = row_fill
            cell.alignment = WRAP_ALIGN if col <= 2 else CENTER_ALIGN


def write_table_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[list],
    footer_lines: list[str] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    header_row = 3
    data_start = header_row + 1

    for col, h in enumerate(headers, start=1):
        ws.cell(header_row, col, h)

    for r_idx, row in enumerate(rows, start=data_start):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val if val not in ("", "nan") else None)

    style_title_block(ws, header_row)
    style_data_rows(ws, data_start, data_start + len(rows) - 1, len(headers))

    if footer_lines:
        foot = data_start + len(rows) + 1
        for i, line in enumerate(footer_lines):
            ws.merge_cells(start_row=foot + i, start_column=1, end_row=foot + i, end_column=len(headers))
            cell = ws.cell(foot + i, 1, line)
            cell.font = SECTION_FONT if line.endswith(":") or line.startswith("Summary") else META_FONT
            cell.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = ws.cell(data_start, 1)
    autosize_columns(ws)


def write_matrix_sheet(wb: openpyxl.Workbook, matrix: pd.DataFrame) -> None:
    ws = wb.create_sheet("S3_matrix")
    header_row = 3
    data_start = header_row + 1

    ws.row_dimensions[1].height = 25
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, len(matrix.columns)))
    ws.cell(1, 1, TITLE + " — binary method × dataset matrix").font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(wrap_text=True)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(8, len(matrix.columns)))
    ws.cell(2, 1, META + " " + MATRIX_META).font = META_FONT
    ws.cell(2, 1).alignment = Alignment(wrap_text=True)

    headers = [format_accession_display(h) for h in matrix.columns]
    for col, h in enumerate(headers, start=1):
        ws.cell(header_row, col, h)

    for r_idx, row in matrix.iterrows():
        excel_row = data_start + r_idx
        row_fill = WHITE_FILL if r_idx % 2 == 0 else GRAY_FILL
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(excel_row, c_idx, int(val) if c_idx > 1 and str(val).isdigit() else val)
            if c_idx == 1:
                cell.fill = row_fill
                cell.font = Font(bold=False)
                cell.alignment = WRAP_ALIGN
            elif str(val) == "1":
                cell.fill = USED_FILL
                cell.alignment = CENTER_ALIGN
            else:
                cell.fill = row_fill
                cell.alignment = CENTER_ALIGN

    for col in range(1, len(headers) + 1):
        cell = ws.cell(header_row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(headers) + 1):
        letter = get_column_letter(col)
        label = headers[col - 1]
        ws.column_dimensions[letter].width = min(max(len(label) + 2, 11), 36)


def write_accession_key_sheet(wb: openpyxl.Workbook, matrix: pd.DataFrame) -> None:
    rows = []
    for raw in matrix.columns[1:]:
        rows.append(
            [
                raw,
                format_accession_display(raw),
                dataset_type(raw),
                zenodo_resolve_url(raw),
                accession_key_note(raw),
            ]
        )
    write_table_sheet(
        wb,
        "S3_accession_key",
        [
            "Canonical ID (matrix key)",
            "Display label",
            "Repository type",
            "Resolve at (URL)",
            "Notes",
        ],
        rows,
        footer_lines=[
            "Repository-type legend:",
            "• GEO / ArrayExpress / SRA / BioProject: primary omics repository accessions.",
            "• Zenodo (DOI deposit): archival DOI on Zenodo — open via https://doi.org/<DOI>. "
            "Often code, GRNs, processed objects, or paper bundles; not always expression matrices.",
            "• Compendium/Resource: named public resources (e.g. ENCODE, LINCS L1000) without a single accession.",
            ZENODO_NOTE,
        ],
    )


def build_shared(matrix: pd.DataFrame) -> pd.DataFrame:
    shared_rows = []
    for ds in matrix.columns[1:]:
        used = matrix.loc[matrix[ds].astype(str) == "1", "Method"].tolist()
        if len(used) > 1:
            shared_rows.append(
                {
                    "Dataset": format_accession_display(ds),
                    "Type": dataset_type(ds),
                    "Methods (count)": len(used),
                    "Methods": "; ".join(used),
                }
            )
    return pd.DataFrame(shared_rows).sort_values("Methods (count)", ascending=False)


def main() -> None:
    matrix = clean_df(pd.read_csv(DATA / "reprogramming_Clean_Matrix.csv"))
    per_method = clean_df(pd.read_csv(DATA / "DatasetsUsed.csv"))
    shared = build_shared(matrix)

    n_methods = len(matrix)
    n_datasets = len(matrix.columns) - 1
    edges = int(matrix.iloc[:, 1:].astype(int).sum().sum())
    single = sum(1 for ds in matrix.columns[1:] if (matrix[ds].astype(str) == "1").sum() == 1)
    with_data = int((matrix.iloc[:, 1:].astype(int).sum(axis=1) > 0).sum())

    footer = [
        "Summary Statistics:",
        f"• Methods with ≥1 catalogued validation dataset: {with_data} / {n_methods} ({100 * with_data / n_methods:.1f}%)",
        f"• Total unique datasets catalogued: {n_datasets}",
        f"• Method–dataset edges: {edges}",
        f"• Datasets used by exactly one method: {single} ({100 * single / n_datasets:.1f}%)",
        f"• Datasets shared across methods: {len(shared)} (see sheet S3_shared)",
        "Repository types and Zenodo DOIs:",
        "• Catalog includes GEO/SRA/BioProject accessions, named compendia (ENCODE, LINCS L1000, …), "
        "and Zenodo DOIs formatted as DOI 10.5281/zenodo.<record> (open at https://doi.org/…).",
        f"• {ZENODO_NOTE}",
        "• Full accession list with repository types: sheet S3_accession_key.",
        "Column definitions:",
        "• Datasets Used: semicolon-separated accessions from manual curation and programmatic extraction.",
        "• Empty Datasets Used: no external accession reported, bundled/compendium-only, or model-based method.",
    ]

    per_headers = [
        "#",
        "Method Name",
        "PubMedID",
        "Year",
        "Type of Reprogramming",
        "Datasets Used",
        "N datasets",
    ]
    per_rows = []
    for i, row in per_method.iterrows():
        ds = format_datasets_used_list(str(row.get("Datasets Used", "") or "").strip())
        n = len([x for x in re.split(r"[;,]", ds) if x.strip()]) if ds else 0
        per_rows.append(
            [
                i + 1,
                row.get("Method Name", ""),
                row.get("PubMedID", ""),
                row.get("Year", ""),
                row.get("Type of Reprogramming", ""),
                ds,
                n if n else "",
            ]
        )

    shared_headers = ["Dataset", "Type", "Methods (count)", "Methods"]
    shared_rows = [list(r) for r in shared.itertuples(index=False, name=None)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_table_sheet(wb, "S3 Table", per_headers, per_rows, footer_lines=footer)
    write_table_sheet(
        wb,
        "S3_shared",
        shared_headers,
        shared_rows,
        footer_lines=[
            "Datasets utilized by more than one method (derived from S3_matrix). "
            "Zenodo entries are shown as DOI 10.5281/zenodo.<record>.",
        ],
    )
    write_accession_key_sheet(wb, matrix)
    write_matrix_sheet(wb, matrix)

    wb.save(OUT)
    shutil.copy2(OUT, OUT_LEGACY)
    print(f"Wrote {OUT}")
    print(f"Copied legacy mirror → {OUT_LEGACY}")


if __name__ == "__main__":
    main()
